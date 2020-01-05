# import openpyxl and tkinter modules
from openpyxl import *
from tkinter import *

# globally declare wb and sheet variable

# opening the existing excel file
wb = load_workbook('C:\\Users\\devvr\\OSTL_Project\\excel.xlsx')

# create the sheet object
sheet = wb.active

def excel():
    # resize the width of columns in
    # excel spreadsheet
    sheet.column_dimensions['A'].width = 30
    sheet.column_dimensions['B'].width = 10
    sheet.column_dimensions['C'].width = 10
    sheet.column_dimensions['D'].width = 20
    sheet.column_dimensions['E'].width = 20
    sheet.column_dimensions['F'].width = 20
    sheet.column_dimensions['G'].width = 20
    sheet.column_dimensions['H'].width = 20
    sheet.column_dimensions['I'].width = 20
    sheet.column_dimensions['J'].width = 30

    # write given data to an excel spreadsheet
    # at particular location
    sheet.cell(row=1, column=1).value = "Name"
    sheet.cell(row=1, column=2).value = "Gender"
    sheet.cell(row=1, column=3).value = "Course"
    sheet.cell(row=1, column=4).value = "Semester"
    sheet.cell(row=1, column=5).value = "No. of Languages"
    sheet.cell(row=1, column=6).value = "First Language Learnt"
    sheet.cell(row=1, column=7).value = "Prefered Programming Language"
    sheet.cell(row=1, column=8).value = "After Degree"
    sheet.cell(row=1, column=9).value = "Contact Number"
    sheet.cell(row=1, column=10).value = "Email id"


# Function to set focus(cursor)
def focus1(event):
    # set focus on the gender_fieldbx
    gender_field.focus_set()


# Function to set focus (cursor)
def focus2(event):
    # set focus on the course_field box
    course_field.focus_set()


# Function to set focus
def focus3(event):
    # set focus on the sem_field box
    sem_field.focus_set()


# Function to set focus
def focus4(event):
    # set focus on the  program1_field box
    program1_field.focus_set()


# Function to set focus
def focus5(event):
    # set focus on the first_language_field box
    first_language_field.focus_set()


# Function to set focus
def focus6(event):
    # set focus on the preferred_lan_field box
    prefered_lan_field.focus_set()


# Function to set focus
def focus7(event):
    # set focus on the after_degree_field box
    af_degree_field.focus_set()


# Function to set focus
def focus8(event):
    # set focus on the contact_no_field box
    contact_no_field.focus_set()


# Function to set focus
def focus9(event):
    # set focus on the email_id_field box
    email_id_field.focus_set()


# Function for clearing the
# contents of text entry boxes
def clear():
    # clear the content of text entry box
    name_field.delete(0, END)
    course_field.delete(0, END)
    sem_field.delete(0, END)
    first_language_field.delete(0, END)
    prefered_lan_field.delete(0, END)
    contact_no_field.delete(0, END)
    email_id_field.delete(0, END)


# Function to take data from GUI
# window and write to an excel file
def insert():
    global v
    # if user not fill any entry
    # then print "empty input"
    if (name_field.get() == "" and
            course_field.get() == "" and
            sem_field.get() == "" and
            first_language_field.get() == "" and
            prefered_lan_field.get() == "" and
            contact_no_field.get() == "" and
            email_id_field.get() == " "):

        print("Empty input")

    else:
        # assigning the max row and max column
        # value upto which data is written
        # in an excel sheet to the variable
        current_row = sheet.max_row
        current_column = sheet.max_column

        # get method returns current text
        # as string which we write into
        # excel spreadsheet at particular location
        sheet.cell(row=current_row + 1, column=1).value = name_field.get()
        sheet.cell(row=current_row + 1, column=2).value = v.get()
        sheet.cell(row=current_row + 1, column=3).value = course_field.get()
        sheet.cell(row=current_row + 1, column=4).value = sem_field.get()
        sheet.cell(row=current_row + 1, column=5).value = v1.get()
        sheet.cell(row=current_row + 1, column=6).value = first_language_field.get()
        sheet.cell(row=current_row + 1, column=7).value = prefered_lan_field.get()
        sheet.cell(row=current_row + 1, column=8).value = v2.get()
        sheet.cell(row=current_row + 1, column=9).value = contact_no_field.get()
        sheet.cell(row=current_row + 1, column=10).value = email_id_field.get()

        # save the file
        wb.save('C:\\Users\\devvr\\OSTL_Project\\excel.xlsx')

        # set focus on the name_field box
        name_field.focus_set()

        # call the clear() function
        clear()

    # Driver code


if __name__ == "__main__":
    # create a GUI window
    root = Tk()
    # set the background colour of GUI window
    root.configure(background='#38ACEC')
    # set the title of GUI window
    root.title("registration form")
    # set the configuration of GUI window
    root.geometry("700x400")
    excel()

    # create a Form label
    heading = Label(root, text="Form", font="Verdana 20 bold", bg="#38ACEC")

    # create a Name label
    name = Label(root, text="Name", bg="#38ACEC")

    # create male / female label
    gender = Label(root, text="Gender", bg="#38ACEC")
    # radiobuttons
    v = StringVar()
    value = ""
    radio1 = Radiobutton(root, text="Male", variable=v, value="Male")
    radio2 = Radiobutton(root, text="Female", variable=v, value="Female")

    # create a Course label
    course = Label(root, text="Course", bg="#38ACEC")

    # create a Semester label
    sem = Label(root, text="Semester", bg="#38ACEC")

    # No. of programming languages
    program1 = Label(root, text="No. of Programming languages learnt?", bg="#38ACEC")
    # radiobuttons
    v1 = IntVar()
    value = 0
    radio3 = Radiobutton(root, text="1", variable=v1, value=1)
    radio4 = Radiobutton(root, text="2", variable=v1, value=2)
    radio5 = Radiobutton(root, text="3", variable=v1, value=3)
    radio6 = Radiobutton(root, text="4", variable=v1, value=4)
    radio7 = Radiobutton(root, text="5", variable=v1, value=5)

    # First Language learnt
    first_language = Label(root, text="First Programming Language", bg="#38ACEC")

    # create preferred programming languages
    prefered_lan = Label(root, text="Prefered Language?", bg="#38ACEC")

    # create a After Degree things
    af_degree = Label(root, text="After BE, what will be your choice?", bg="#38ACEC")

    # radiobuttons
    v2 = StringVar()
    value = " "
    radio8 = Radiobutton(root, text="Job", variable=v2, value="Job")
    radio9 = Radiobutton(root, text="ME/M-Tech", variable=v2, value="ME/M-Tech")
    radio10 = Radiobutton(root, text="MS", variable=v2, value="MS")
    radio11 = Radiobutton(root, text="Business", variable=v2, value="Business")
    radio12 = Radiobutton(root, text="MBA", variable=v2, value="MBA")
    radio13 = Radiobutton(root, text="Entrepreneur", variable=v2, value="Entrepreneur")

    # create a Contact No. label
    contact_no = Label(root, text="Contact No.", bg="#38ACEC")

    # create a Email-id label
    email_id = Label(root, text="Email id", bg="#38ACEC")

    # grid method is used for placing
    # the widgets at respective positions
    # in table like structure .
    heading.grid(row=0, column=1)
    name.grid(row=1, column=0)
    gender.grid(row=2, column=0)
    course.grid(row=3, column=0)
    sem.grid(row=4, column=0)
    program1.grid(row=5, column=0)
    first_language.grid(row=8, column=0)
    prefered_lan.grid(row=9, column=0)
    af_degree.grid(row=10, column=0)
    contact_no.grid(row=13, column=0)
    email_id.grid(row=14, column=0)

    # create a text entry box
    # for typing the information
    name_field = Entry(root)
    gender_field = Entry(root)
    course_field = Entry(root)
    sem_field = Entry(root)
    program1_field = Entry(root)
    first_language_field = Entry(root)
    prefered_lan_field = Entry(root)
    af_degree_field = Entry(root)
    contact_no_field = Entry(root)
    email_id_field = Entry(root)

    # bind method of widget is used for
    # the binding the function with the events
    # whenever the enter key is pressed
    # then call the focus1 function
    name_field.bind("<Return>", focus1)

    # whenever the enter key is pressed
    # then call the focus2 function
    gender_field.bind("<Return>", focus2)

    # whenever the enter key is pressed
    # then call the focus3 function
    course_field.bind("<Return>", focus3)

    # whenever the enter key is pressed
    # then call the focus4 function
    sem_field.bind("<Return>", focus4)

    # whenever the enter key is pressed
    # then call the focus5 function
    program1_field.bind("<Return>", focus5)

    # whenever the enter key is pressed
    # then call the focus6 function
    first_language_field.bind("<Return>", focus6)

    # whenever the enter key is pressed
    # then call the focus7 function
    prefered_lan_field.bind("<Return>", focus7)

    # whenever the enter key is pressed
    # then call the focus8 function
    af_degree_field.bind("<Return>", focus8)

    # whenever the enter key is pressed
    # then call the focus9 function
    contact_no_field.bind("<Return>", focus9)

    # whenever the enter key is pressed
    # then call the focus9 function
    email_id_field.bind("<Return>", focus9)

    # grid method is used for placing
    # the widgets at respective positions
    # in table like structure .
    name_field.grid(row=1, column=1, ipadx="100")
    # gender_field.grid(row = 2, column = 1, ipadx = "100")
    radio1.grid(row=2, column=1, sticky=W)
    radio2.grid(row=2, column=1)
    course_field.grid(row=3, column=1, ipadx="100")
    sem_field.grid(row=4, column=1, ipadx="100")
    radio3.grid(row=5, column=1, sticky=W)
    radio4.grid(row=5, column=1)
    radio5.grid(row=6, column=1, sticky=W)
    radio6.grid(row=6, column=1)
    radio7.grid(row=7, column=1, sticky=W)
    first_language_field.grid(row=8, column=1, ipadx="100")
    prefered_lan_field.grid(row=9, column=1, ipadx="100")
    radio8.grid(row=10, column=1, sticky=W)
    radio9.grid(row=10, column=1)
    radio10.grid(row=11, column=1, sticky=W)
    radio11.grid(row=11, column=1)
    radio12.grid(row=12, column=1, sticky=W)
    radio13.grid(row=12, column=1)
    contact_no_field.grid(row=13, column=1, ipadx="100")
    email_id_field.grid(row=14, column=1, ipadx="100")

    # call excel function
    excel()

    # create a Submit Button and place into the root window
    submit = Button(root, text="Submit", fg="Black", bg="White", font="Arial 10 bold", command=insert)
    submit.grid(row=15, column=1)

    # start the GUI
    root.mainloop()
